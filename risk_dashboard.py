"""
Risk Assessment Dashboard - Streamlit + AgGrid

Applicazione web per la gestione completa dei rischi di progetto con:
- Interfaccia interattiva per l'aggiunta di nuovi rischi
- Tabella dinamica con funzionalità di modifica ed eliminazione
- Heat map visuale per analisi dei rischi
- Esportazione in PDF ed Excel con report dettagliati

Lancia con:
    streamlit run risk_dashboard.py

Dipendenze:
    pip install streamlit streamlit-aggrid pandas reportlab matplotlib openpyxl pillow

Autore: [Nome]
Data: [Data]
Versione: 1.0
"""

# ===========================
# IMPORTAZIONI E CONFIGURAZIONI
# ===========================

import streamlit as st
import pandas as pd
import os
from datetime import date
from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode, JsCode, DataReturnMode
from io import BytesIO

# ===========================
# CONFIGURAZIONI GLOBALI
# ===========================

# Nome del file CSV per la persistenza dei dati
DATA_FILE = 'risk_data.csv'

# Configurazione layout Streamlit per utilizzo completo della larghezza
st.set_page_config(page_title="Dashboard Risk Assessment", layout="wide")

# Caricamento stili CSS personalizzati
with open("risk_dashboard_styles.css") as f:
    st.markdown(f"<style>{f.read()}</style>", unsafe_allow_html=True)

# ===========================
# HEADER E TITOLO PRINCIPALE
# ===========================

st.markdown("""
<div style='width:100%; text-align:center; margin-bottom: 40px;'>
    <h1 style='
        display: inline-block; 
        margin: 0; 
        color: #e0e6f0;
        background: linear-gradient(135deg, #8B0000 0%, #DC143C 50%, #FF6347 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        background-clip: text;
        font-size: 3rem;
        font-weight: 700;
        text-shadow: 0 4px 8px rgba(220, 20, 60, 0.3);
    '>
        🛡️ Dashboard Risk Assessment
    </h1>
    <div style='
        width: 100px; 
        height: 4px; 
        background: linear-gradient(135deg, #8B0000 0%, #DC143C 50%, #FF6347 100%); 
        margin: 16px auto;
        border-radius: 2px;
    '></div>
</div>
""", unsafe_allow_html=True)

# ===========================
# FUNZIONI DI GESTIONE DATI
# ===========================

def load_data(file_path):
    """
    Carica i dati dei rischi dal file CSV specificato.
    
    Args:
        file_path (str): Percorso del file CSV contenente i dati dei rischi
        
    Returns:
        pd.DataFrame: DataFrame contenente i dati dei rischi con tipizzazione corretta
                     delle colonne numeriche. Ritorna DataFrame vuoto se file non esiste.
                     
    Raises:
        Exception: Gestisce errori di lettura file mostrando messaggio di errore
    """
    if os.path.exists(file_path):
        try:
            df = pd.read_csv(file_path)
            # Conversione e validazione dei tipi di dati per garantire consistenza
            if not df.empty:
                df['ID'] = df['ID'].astype(int)
                df['Probabilità'] = df['Probabilità'].astype(float)  # Supporta valori decimali (0.5)
                df['Impatto'] = df['Impatto'].astype(float)        # Supporta valori decimali (0.5)
                df['Valore_Rischio'] = df['Valore_Rischio'].astype(float)  # Prodotto prob*impatto
            return df
        except Exception as e:
            st.error(f"Errore nel caricamento dei dati: {e}")
            return create_empty_dataframe()
    else:
        return create_empty_dataframe()

def create_empty_dataframe():
    """
    Crea un DataFrame vuoto con la struttura standard per i dati dei rischi.
    
    Returns:
        pd.DataFrame: DataFrame vuoto con colonne predefinite per i rischi
    """
    return pd.DataFrame({
        "ID": [],
        "Descrizione": [],
        "Probabilità": [],
        "Impatto": [],
        "Valore_Rischio": [],  # Calcolato automaticamente (Probabilità * Impatto)
        "Priorità": [],        # Derivata dal Valore_Rischio
        "Contromisura": [],
        "Stato": [],
        "Data scadenza": []
    })

def save_data(df, file_path):
    """
    Salva il DataFrame dei rischi nel file CSV specificato.
    
    Args:
        df (pd.DataFrame): DataFrame contenente i dati dei rischi da salvare
        file_path (str): Percorso di destinazione per il file CSV
        
    Returns:
        bool: True se il salvataggio è riuscito, False altrimenti
    """
    try:
        df.to_csv(file_path, index=False)
        return True
    except Exception as e:
        st.error(f"Errore nel salvataggio dei dati: {e}")
        return False

# ===========================
# INIZIALIZZAZIONE SESSION STATE
# ===========================

# Inizializzazione dei dati principali
if 'df' not in st.session_state:
    st.session_state.df = load_data(DATA_FILE)

def refresh_data():
    """Forza il ricaricamento dei dati dal file per garantire sincronizzazione."""
    st.session_state.df = load_data(DATA_FILE)

# Gestione refresh pagina per sincronizzazione dati
if 'page_refresh' not in st.session_state:
    st.session_state.page_refresh = 0
    refresh_data()

# ===========================
# FUNZIONI DI BUSINESS LOGIC
# ===========================

def calcola_priorita(valore_rischio):
    """
    Calcola la priorità del rischio basata sul valore numerico del rischio.
    
    La scala utilizza il prodotto Probabilità × Impatto (range 1-25):
    - Bassa: 1-5
    - Media: 6-10  
    - Alta: 11-15
    - Estrema: 16-25
    
    Args:
        valore_rischio (float): Valore numerico del rischio (Probabilità × Impatto)
        
    Returns:
        str: Livello di priorità del rischio
    """
    if valore_rischio >= 16:
        return "Estrema"
    elif valore_rischio >= 11:
        return "Alta"
    elif valore_rischio >= 6:
        return "Media"
    else:
        return "Bassa"

def elimina_rischio(risk_id):
    """
    Elimina un rischio specifico dal dataset e salva immediatamente.
    
    Args:
        risk_id (int): ID univoco del rischio da eliminare
        
    Returns:
        bool: True se l'eliminazione e il salvataggio sono riusciti
    """
    # Filtra i dati rimuovendo il rischio con ID specificato
    st.session_state.df = st.session_state.df[st.session_state.df['ID'] != risk_id]
    
    # Persistenza immediata per evitare perdita dati
    if save_data(st.session_state.df, DATA_FILE):
        return True
    return False

# ===========================
# FUNZIONI DI REPORTISTICA
# ===========================

def create_heatmap_image(df):
    """
    Genera una heat map dei rischi come immagine PNG per l'inclusione nei report PDF.
    
    La heat map visualizza:
    - Griglia 5x5 rappresentando combinazioni Probabilità/Impatto
    - Colori basati sui livelli di priorità
    - Posizionamento preciso dei rischi basato sui valori decimali
    - Etichette con ID dei rischi per identificazione
    
    Args:
        df (pd.DataFrame): DataFrame contenente i dati dei rischi
        
    Returns:
        BytesIO: Buffer contenente l'immagine PNG della heat map,
                None se matplotlib non è disponibile o si verifica un errore
    """
    try:
        import matplotlib.pyplot as plt
        import matplotlib.patches as patches
        import numpy as np
        from io import BytesIO
        
        # Configurazione matplotlib per ambiente headless
        plt.switch_backend('Agg')
        
        # Inizializzazione figura con sfondo bianco per PDF
        fig, ax = plt.subplots(figsize=(10, 8))
        fig.patch.set_facecolor('white')
        
        def get_risk_color(prob, imp):
            """Determina il colore della cella basato sul valore del rischio."""
            risk_value = prob * imp
            if risk_value >= 16:
                return '#ef4444'  # Rosso - Rischio Estremo
            elif risk_value >= 11:
                return '#f97316'  # Arancione - Rischio Alto
            elif risk_value >= 6:
                return '#eab308'  # Giallo - Rischio Medio
            elif risk_value >= 1:
                return '#22c55e'  # Verde - Rischio Basso
            else:
                return '#9ca3af'  # Grigio - Nessun rischio
        
        # Creazione griglia di sfondo 5x5
        for i in range(5):  # Righe: Impatto (1-5, dal basso all'alto)
            for j in range(5):  # Colonne: Probabilità (1-5, da sinistra a destra)
                prob = j + 1  
                imp = i + 1   
                color = get_risk_color(prob, imp)
                
                # Disegno rettangolo con colore appropriato
                rect = patches.Rectangle((j, i), 1, 1, linewidth=2, 
                                       edgecolor='#334155', facecolor=color, alpha=0.7)
                ax.add_patch(rect)
        
        # Raggruppamento rischi per posizione per gestire sovrapposizioni
        risk_positions = {}
        for _, row in df.iterrows():
            prob = float(row['Probabilità'])
            imp = float(row['Impatto'])
            risk_id = int(row['ID'])
            position_key = (prob, imp)
            
            if position_key not in risk_positions:
                risk_positions[position_key] = []
            risk_positions[position_key].append(risk_id)
        
        # Posizionamento marcatori dei rischi
        for (prob, imp), risk_ids in risk_positions.items():
            if prob >= 1 and imp >= 1:  # Validazione range valori
                # Calcolo posizione precisa basata su valori decimali
                x = prob - 0.5  # Conversione da scala 1-5 a coordinate 0.5-4.5
                y = imp - 0.5   
                
                # Creazione etichetta con ID multipli se necessario
                id_string = ', '.join(map(str, sorted(risk_ids)))
                
                # Dimensionamento dinamico basato su numero di rischi
                size = 200 + len(risk_ids) * 100
                
                # Aggiunta marcatore visuale
                ax.scatter(x, y, s=size, c='black', edgecolors='white', 
                          linewidth=2, zorder=10)
                
                # Etichetta testuale con ID
                ax.text(x, y, id_string, ha='center', va='center', 
                       color='white', fontweight='bold', fontsize=8, zorder=11)
        
        # Configurazione assi e layout
        ax.set_xlim(0, 5)
        ax.set_ylim(0, 5)
        ax.set_aspect('equal')
        
        # Etichettatura assi
        prob_labels = ['1', '2', '3', '4', '5']
        imp_labels = ['1', '2', '3', '4', '5']
        
        ax.set_xticks([0.5, 1.5, 2.5, 3.5, 4.5])
        ax.set_xticklabels(prob_labels, fontsize=9)
        ax.set_yticks([0.5, 1.5, 2.5, 3.5, 4.5])
        ax.set_yticklabels(imp_labels, fontsize=9)
        
        # Titoli e etichette
        ax.set_xlabel('PROBABILITÀ', fontweight='bold', fontsize=12, labelpad=10)
        ax.set_ylabel('IMPATTO', fontweight='bold', fontsize=12, labelpad=10)
        ax.set_title('HEAT MAP DEI RISCHI', fontweight='bold', fontsize=14, pad=20)
        
        # Griglia di supporto
        ax.grid(True, alpha=0.3)
        ax.set_axisbelow(True)
        
        # Legenda colori per priorità
        legend_elements = [
            patches.Patch(color='#22c55e', label='Bassa (1-5)'),
            patches.Patch(color='#eab308', label='Media (6-10)'),
            patches.Patch(color='#f97316', label='Alta (11-15)'),
            patches.Patch(color='#ef4444', label='Estrema (16-25)')
        ]
        ax.legend(handles=legend_elements, loc='upper left', bbox_to_anchor=(1.02, 1), 
                 title='Priorità', title_fontsize=10, fontsize=9)
        
        # Salvataggio in buffer per utilizzo nei report
        img_buffer = BytesIO()
        plt.tight_layout()
        plt.savefig(img_buffer, format='PNG', dpi=300, bbox_inches='tight', 
                   facecolor='white', edgecolor='none')
        plt.close()
        
        img_buffer.seek(0)
        return img_buffer
        
    except ImportError:
        st.warning("Matplotlib non disponibile. Heat map non inclusa nel PDF.")
        return None
    except Exception as e:
        st.warning(f"Errore nella creazione della heat map: {str(e)}")
        return None

def create_pdf_report(df):
    """
    Genera un report PDF completo contenente tabella dei rischi, riepilogo e heat map.
    
    Il report include:
    - Tabella formattata con tutti i rischi
    - Riepilogo statistico per priorità
    - Heat map visuale dei rischi
    - Formattazione professionale con intestazioni e stili
    
    Args:
        df (pd.DataFrame): DataFrame contenente i dati dei rischi
        
    Returns:
        BytesIO: Buffer contenente il PDF generato,
                None se le librerie non sono disponibili o si verifica un errore
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib.enums import TA_CENTER
        import io

        # Configurazione documento in formato landscape per tabelle larghe
        pdf_buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            pdf_buffer,
            pagesize=landscape(A4),
            rightMargin=20,
            leftMargin=20,
            topMargin=20,
            bottomMargin=20
        )

        elements = []
        styles = getSampleStyleSheet()
        
        # Stile personalizzato per il titolo principale
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=20,
            spaceAfter=20,
            alignment=TA_CENTER
        )

        # Intestazione del documento
        elements.append(Paragraph("🛡️ Risk Assessment Report", title_style))
        elements.append(Paragraph(f"Data: {date.today().strftime('%d/%m/%Y')}", styles['Normal']))
        elements.append(Spacer(1, 0.3 * inch))

        # Preparazione dati tabella con intestazioni in ordine coerente con interfaccia AgGrid
        data = [['ID', 'Descrizione', 'Probabilità', 'Impatto', 'Priorità', 'Contromisura', 'Stato', 'Data Scadenza']]

        # Popolamento righe dati con text wrapping intelligente e formattazione ottimizzata
        for _, row in df.iterrows():
            # Descrizione con wrapping intelligente (max 50 caratteri per riga PDF)
            desc = str(row['Descrizione']) if pd.notnull(row['Descrizione']) else ""
            if len(desc) > 50:
                # Spezza a circa 50 caratteri mantenendo parole intere
                words = desc.split(' ')
                lines = []
                current_line = ""
                for word in words:
                    if len(current_line + " " + word) <= 50:
                        current_line += (" " + word) if current_line else word
                    else:
                        if current_line:
                            lines.append(current_line)
                        current_line = word
                if current_line:
                    lines.append(current_line)
                desc = "\n".join(lines[:2])  # Max 2 righe per PDF
                if len(lines) > 2:
                    desc += "..."
            
            # Contromisura con wrapping simile
            desc_cont = str(row['Contromisura']) if pd.notnull(row['Contromisura']) else ""
            if len(desc_cont) > 45:
                words = desc_cont.split(' ')
                lines = []
                current_line = ""
                for word in words:
                    if len(current_line + " " + word) <= 45:
                        current_line += (" " + word) if current_line else word
                    else:
                        if current_line:
                            lines.append(current_line)
                        current_line = word
                if current_line:
                    lines.append(current_line)
                desc_cont = "\n".join(lines[:2])  # Max 2 righe
                if len(lines) > 2:
                    desc_cont += "..."

            # Aggiunta riga con ordine colonne coerente con AgGrid
            data.append([
                str(row['ID']),
                desc,
                f"{row['Probabilità']:.1f}",
                f"{row['Impatto']:.1f}",
                str(row['Priorità']),
                desc_cont,
                str(row['Stato']),
                str(row['Data scadenza'])  # Aggiunta colonna mancante
            ])

        # Creazione e styling avanzato della tabella con larghezze ridotte per eliminare spazio vuoto
        table = Table(data, colWidths=[
            0.5*inch,   # ID - compatta (coerente con 60px AgGrid)
            2.8*inch,   # Descrizione - ottimizzata (coerente con 300px AgGrid)
            0.7*inch,   # Probabilità (coerente con 100px AgGrid)
            0.6*inch,   # Impatto (coerente con 90px AgGrid)
            0.7*inch,   # Priorità (coerente con 100px AgGrid)
            2.6*inch,   # Contromisura - ottimizzata (coerente con 280px AgGrid)
            0.7*inch,   # Stato (coerente con 80px AgGrid)
            1.0*inch    # Data Scadenza (coerente con 110px AgGrid)
        ])
        
        table_style = TableStyle([
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),      
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke), 
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),             
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),   
            ('FONTSIZE', (0, 0), (-1, 0), 9),                  # Font header ridotto per spazio
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),            
            
            # Contenuto styling
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),       
            ('FONTSIZE', (0, 1), (-1, -1), 8),                 # Font contenuto ridotto
            ('GRID', (0, 0), (-1, -1), 1, colors.black),       
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),               # Allineamento top per text wrapping
            
            # Allineamento specifico per colonne di testo
            ('ALIGN', (1, 1), (1, -1), 'LEFT'),               # Descrizione a sinistra
            ('ALIGN', (5, 1), (5, -1), 'LEFT'),               # Contromisura a sinistra
            
            # Padding ottimizzato per text wrapping
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ])
        table.setStyle(table_style)
        elements.append(table)

        # Sezione riepilogo statistico
        elements.append(Spacer(1, 0.5 * inch))
        elements.append(Paragraph("Riepilogo per Priorità", styles['Heading2']))
        priority_counts = df['Priorità'].value_counts()
        summary_text = ""
        for priority in ['Estrema', 'Alta', 'Media', 'Bassa']:
            count = priority_counts.get(priority, 0)
            summary_text += f"{priority}: {count} rischi<br/>"
        elements.append(Paragraph(summary_text, styles['Normal']))

        # Aggiunta heat map se disponibile
        heatmap_img = create_heatmap_image(df)
        if heatmap_img is not None:
            elements.append(PageBreak())
            elements.append(Paragraph("Heat Map dei Rischi", styles['Heading2']))
            elements.append(Spacer(1, 0.2 * inch))
            img = Image(heatmap_img, width=7 * inch, height=5 * inch)
            elements.append(img)
            elements.append(Spacer(1, 0.3 * inch))
            elements.append(Paragraph(
                "Note: I numeri neri indicano gli ID dei rischi posizionati secondo probabilità e impatto (valori da 1 a 5 con incrementi di 0.5). "
                "I colori rappresentano le priorità: Verde (Bassa), Giallo (Media), Arancione (Alta), Rosso (Estrema).",
                styles['Normal']))

        # Generazione finale del PDF
        doc.build(elements)
        pdf_buffer.seek(0)
        return pdf_buffer

    except ImportError:
        st.error("Librerie necessarie non disponibili. Installa con: pip install reportlab matplotlib")
        return None
    except Exception as e:
        st.error(f"Errore nella creazione del PDF: {str(e)}")
        return None

# ===========================
# INIZIALIZZAZIONE FORM STATE
# ===========================

# Inizializzazione valori form per gestione stato persistente tra submit
form_defaults = {
    'form_descrizione': "",
    'form_prob': 1.0,
    'form_imp': 1.0,
    'form_contromisura': "",
    'form_stato': "Da pianificare",
    'form_data_scad': date.today(),
    'form_counter': 0  # Counter per forzare reset form dopo submit
}

for key, default_value in form_defaults.items():
    if key not in st.session_state:
        st.session_state[key] = default_value

# ===========================
# INTERFACCIA FORM AGGIUNTA RISCHI
# ===========================

st.header("Aggiungi Nuovo Rischio")

# Form con chiave dinamica per permettere reset dopo submit successful
with st.form(key=f'form_nuovo_rischio_{st.session_state.form_counter}'):
    # Input descrizione del rischio
    descrizione = st.text_input(
        "Descrizione del rischio", 
        value=st.session_state.form_descrizione,
        help="Descrizione dettagliata del rischio identificato"
    )
    
    # Slider per probabilità con incrementi di 0.5 per granularità fine
    prob = st.slider(
        "Probabilità (1-5)", 
        min_value=1.0, 
        max_value=5.0, 
        value=st.session_state.form_prob, 
        step=0.5,
        help="Probabilità che il rischio si verifichi (1=Molto Bassa, 5=Molto Alta)"
    )
    
    # Slider per impatto con incrementi di 0.5
    imp = st.slider(
        "Impatto (1-5)", 
        min_value=1.0, 
        max_value=5.0, 
        value=st.session_state.form_imp, 
        step=0.5,
        help="Impatto del rischio se si verifica (1=Minimo, 5=Catastrofico)"
    )
    
    # Input contromisura preventiva/correttiva
    contromisura = st.text_input(
        "Contromisura", 
        value=st.session_state.form_contromisura,
        help="Azioni preventive o correttive per mitigare il rischio"
    )
    
    # Selectbox per stato del rischio
    stato = st.selectbox(
        "Stato", 
        options=["Da pianificare", "In corso", "Monitoraggio", "Chiuso"], 
        index=["Da pianificare", "In corso", "Monitoraggio", "Chiuso"].index(st.session_state.form_stato),
        help="Stato attuale della gestione del rischio"
    )
    
    # Date picker per scadenza azioni
    data_scad = st.date_input(
        "Data scadenza", 
        value=st.session_state.form_data_scad,
        help="Data entro cui completare le azioni di mitigazione"
    )
    
    # Submit button del form
    submit_button = st.form_submit_button(label='Aggiungi')
    
    # Logica di elaborazione submit
    if submit_button:
        if descrizione.strip() != '':
            # Calcolo automatico valore e priorità del rischio
            valore = prob * imp
            priorita = calcola_priorita(valore)
            data_scad_str = data_scad.strftime("%Y-%m-%d")
            
            # Generazione ID incrementale automatico
            nuovo_id = int(st.session_state.df["ID"].max()) + 1 if not st.session_state.df.empty else 1
            
            # Creazione nuovo record rischio
            new_row = {
                "ID": nuovo_id,
                "Descrizione": descrizione,
                "Probabilità": prob,
                "Impatto": imp,
                "Valore_Rischio": valore,
                "Priorità": priorita,
                "Contromisura": contromisura,
                "Stato": stato,
                "Data scadenza": data_scad_str
            }
            
            # Aggiunta al DataFrame principale
            new_df = pd.DataFrame([new_row])
            st.session_state.df = pd.concat([st.session_state.df, new_df], ignore_index=True)
            
            # Persistenza dati e reset form in caso di successo
            if save_data(st.session_state.df, DATA_FILE):
                # Reset automatico form per facilitare inserimenti multipli
                st.session_state.form_descrizione = ""
                st.session_state.form_prob = 1.0
                st.session_state.form_imp = 1.0
                st.session_state.form_contromisura = ""
                st.session_state.form_stato = "Da pianificare"
                st.session_state.form_data_scad = date.today()
                st.session_state.form_counter += 1  # Incremento per forzare ricreazione form
                
                st.success("Nuovo rischio aggiunto!")
                st.rerun()
            else:
                st.error("Errore nel salvataggio del nuovo rischio")
        else:
            st.warning("La descrizione non può essere vuota.")

# ===========================
# JAVASCRIPT PER UX MIGLIORATA
# ===========================

# Script per gestione tasto Invio nel form per submit rapido
st.markdown("""
<script>
document.addEventListener('DOMContentLoaded', function() {
    // Trova tutti gli input di testo e data nel form
    const formInputs = document.querySelectorAll('input[type="text"], input[type="date"]');
    
    // Aggiunge listener per tasto Invio su ogni input
    formInputs.forEach(function(input) {
        input.addEventListener('keydown', function(event) {
            if (event.key === 'Enter') {
                event.preventDefault();
                // Trova e attiva il pulsante submit del form
                const submitButton = document.querySelector('button[kind="formSubmit"]');
                if (submitButton) {
                    submitButton.click();
                }
            }
        });
    });
});

// Listener globale alternativo per cattura Invio universale nel form
document.addEventListener('keydown', function(event) {
    if (event.key === 'Enter') {
        // Verifica se l'elemento attivo è all'interno di un form
        const activeElement = document.activeElement;
        const isInForm = activeElement && (
            activeElement.closest('form') || 
            activeElement.tagName === 'INPUT' || 
            activeElement.tagName === 'TEXTAREA'
        );
        
        if (isInForm) {
            const submitButton = document.querySelector('button[kind="formSubmit"]');
            if (submitButton) {
                event.preventDefault();
                submitButton.click();
            }
        }
    }
});
</script>
""", unsafe_allow_html=True)

# ===========================
# INTERFACCIA TABELLA RISCHI
# ===========================

st.header("Tabella dei rischi")
st.markdown("<div style='margin-bottom: 16px;'></div>", unsafe_allow_html=True)

# Elaborazione e visualizzazione tabella solo se ci sono dati
if not st.session_state.df.empty:
    # Preparazione DataFrame per visualizzazione AgGrid
    display_df = st.session_state.df.copy()
    display_df = display_df.drop(columns=['Valore_Rischio'])  # Nasconde valore calcolato
    
    # Mappatura icone per stati per migliorare visual feedback
    status_mapping = {
        'Da pianificare': '📋',
        'In corso': '⚡', 
        'Monitoraggio': '👀',
        'Chiuso': '✅'
    }
    display_df['Stato'] = display_df['Stato'].map(status_mapping)
    
    # Aggiunta colonna eliminazione per gestione interattiva
    display_df['Elimina'] = False  # Colonna boolean per checkbox
    
    # ===========================
    # CONFIGURAZIONE AGGRID AVANZATA
    # ===========================
    
    # Inizializzazione builder per configurazione griglia
    gb = GridOptionsBuilder.from_dataframe(display_df)

    # Configurazione globale delle colonne
    gb.configure_default_column(
        groupable=False,           # Disabilita raggruppamento per semplicità
        value=True,               # Abilita visualizzazione valori
        enableRowGroup=False,     # Disabilita row grouping
        aggFunc="sum",           # Funzione aggregazione di default
        editable=False,          # Disabilita editing inline per controllo
        resizable=True,          # Permette ridimensionamento manuale colonne
        wrapText=True,           # Abilita text wrapping per contenuti lunghi
        autoHeight=True          # Altezza automatica basata su contenuto
    )

    # Configurazione specifica colonne con dimensioni fisse e text wrapping
    gb.configure_column("ID", width=60, minWidth=50, maxWidth=80, type=["numericColumn"], header_name="ID",
                    wrapText=True, autoHeight=True,
                    cellStyle={'textAlign': 'center', 'fontSize': '14px', 'fontWeight': '600', 
                                'whiteSpace': 'normal', 'wordWrap': 'break-word'})
    
    gb.configure_column("Descrizione", width=300, minWidth=250, maxWidth=350, header_name="DESCRIZIONE",
                    wrapText=True, autoHeight=True,
                    cellStyle={'textAlign': 'left', 'fontSize': '13px', 'fontWeight': '600',
                                'padding': '8px', 'whiteSpace': 'normal', 'wordWrap': 'break-word', 'lineHeight': '1.4'})
    
    gb.configure_column("Probabilità", width=100, minWidth=80, maxWidth=120, header_name="PROBABILITÀ", 
                    wrapText=True, autoHeight=True,
                    cellStyle={'textAlign': 'center', 'fontSize': '14px', 'fontWeight': '600',
                                'whiteSpace': 'normal', 'wordWrap': 'break-word'})
    
    gb.configure_column("Impatto", width=90, minWidth=70, maxWidth=110, header_name="IMPATTO", 
                    wrapText=True, autoHeight=True,
                    cellStyle={'textAlign': 'center', 'fontSize': '14px', 'fontWeight': '600',
                                'whiteSpace': 'normal', 'wordWrap': 'break-word'})
    
    gb.configure_column("Priorità", width=100, minWidth=80, maxWidth=120, header_name="PRIORITÀ",
                    wrapText=True, autoHeight=True,
                    cellStyle={'textAlign': 'center', 'fontSize': '12px', 'fontWeight': '700',
                                'whiteSpace': 'normal', 'wordWrap': 'break-word'})
    
    gb.configure_column("Contromisura", width=280, minWidth=200, maxWidth=320, header_name="CONTROMISURA",
                    wrapText=True, autoHeight=True,
                    cellStyle={'textAlign': 'left', 'fontSize': '13px', 'fontWeight': '600',
                                'padding': '8px', 'whiteSpace': 'normal', 'wordWrap': 'break-word', 'lineHeight': '1.4'})
    
    gb.configure_column("Stato", width=80, minWidth=60, maxWidth=100, header_name="STATO",
                    wrapText=True, autoHeight=True,
                    cellStyle={'textAlign': 'center', 'fontSize': '16px', 'fontWeight': '500',
                                'whiteSpace': 'normal', 'wordWrap': 'break-word'})
    
    gb.configure_column("Data scadenza", width=110, minWidth=90, maxWidth=130, header_name="SCADENZA", 
                    wrapText=True, autoHeight=True,
                    cellStyle={'textAlign': 'center', 'fontSize': '12px', 'fontWeight': '600',
                                'whiteSpace': 'normal', 'wordWrap': 'break-word'})
    
    # Configurazione colonna eliminazione con editor checkbox
    gb.configure_column("Elimina", width=80, minWidth=60, maxWidth=100, header_name="ELIMINA",
        editable=True,                      # Unica colonna editabile
        cellEditor="agCheckboxCellEditor",  # Editor checkbox nativo AgGrid
        cellRenderer="agCheckboxCellRenderer", # Renderer checkbox
        cellStyle={
            'textAlign': 'center', 
            'height': '60px',
            'display': 'flex',
            'alignItems': 'center',
            'justifyContent': 'center'
            }
        )

    # Configurazioni griglia generali
    gb.configure_selection(selection_mode="single", use_checkbox=False)  # Selezione singola riga
    gb.configure_grid_options(domLayout='normal', getRowHeight=None, autoRowHeight=True)

    # Build configurazione finale
    grid_options = gb.build()

    # ===========================
    # RENDERING AGGRID CON STILI PERSONALIZZATI
    # ===========================

    grid_response = AgGrid(
        display_df,
        gridOptions=grid_options,
        height=400,                                    # Altezza fissa griglia
        width='100%',                                  # Larghezza completa
        data_return_mode=DataReturnMode.FILTERED_AND_SORTED,  # Ritorna dati filtrati/ordinati
        update_mode=GridUpdateMode.VALUE_CHANGED,     # Aggiornamento su modifica valori
        fit_columns_on_grid_load=False,               # Disabilita auto-fit per mantenere dimensioni fisse
        theme='streamlit',                            # Tema base Streamlit
        allow_unsafe_jscode=True,                     # Permette JavaScript personalizzato
        custom_css={
            # Styling toolbar griglia
            "#gridToolBar": {
                "padding-bottom": "0px !important"
            },
            # Container principale griglia con effetti visual premium
            ".ag-theme-streamlit": {
                "border": "2px solid #334155 !important",
                "border-radius": "12px !important",
                "overflow": "hidden !important",
                "box-shadow": "0 8px 32px rgba(0, 0, 0, 0.3) !important"
            },
            # Header con gradient background e styling premium
            ".ag-theme-streamlit .ag-header": {
                "background": "linear-gradient(135deg, #1e293b 0%, #334155 100%) !important",
                "color": "#e2e8f0 !important",
                "font-weight": "600 !important",
                "text-align": "center !important",
                "border-bottom": "2px solid #DC143C !important"
            },
            # Celle header con separatori
            ".ag-theme-streamlit .ag-header-cell": {
                "border-right": "1px solid #475569 !important",
                "text-align": "center !important"
            },
            # Testo header centrato
            ".ag-theme-streamlit .ag-header-cell-text": {
                "text-align": "center !important",
                "width": "100% !important"
            },
            # Righe con tema dark e transizioni smooth
            ".ag-theme-streamlit .ag-row": {
                "background-color": "#0f172a !important",
                "color": "#e2e8f0 !important",
                "border-bottom": "1px solid #1e293b !important",
                "transition": "all 0.2s ease !important"
            },
            # Effetti hover per feedback visuale
            ".ag-theme-streamlit .ag-row:hover": {
                "background-color": "#1e293b !important",
                "transform": "translateY(-1px) !important",
                "box-shadow": "0 4px 12px rgba(220, 20, 60, 0.1) !important"
            },
            # Alternanza colori righe per leggibilità
            ".ag-theme-streamlit .ag-row-even": {
                "background-color": "#1e293b !important"
            },
            ".ag-theme-streamlit .ag-row-odd": {
                "background-color": "#0f172a !important"
            },
            ".ag-theme-streamlit .ag-row-even:hover": {
                "background-color": "#334155 !important"
            },
            ".ag-theme-streamlit .ag-row-odd:hover": {
                "background-color": "#1e293b !important"
            },
            # Stile generale celle con text wrapping forzato
            ".ag-theme-streamlit .ag-cell": {
                "border-right": "1px solid #374151 !important",
                "text-align": "center !important",
                "display": "flex !important",
                "align-items": "center !important",
                "justify-content": "center !important",
                "font-weight": "500 !important",
                "white-space": "normal !important",
                "word-wrap": "break-word !important",
                "word-break": "break-word !important",
                "overflow-wrap": "break-word !important",
                "padding": "6px !important",
                "line-height": "1.3 !important",
                "max-width": "350px !important"
            },
            ".ag-theme-streamlit .ag-cell-value": {
                "text-align": "center !important"
            },
            # Styling speciale per colonna Elimina con gradient di allerta
            ".ag-theme-streamlit [col-id='Elimina'] .ag-cell": {
                "background": "linear-gradient(135deg, rgba(220, 20, 60, 0.1) 0%, rgba(139, 0, 0, 0.15) 100%) !important",
                "border": '2px solid rgba(220, 20, 60, 0.4) !important',
                "border-radius": '8px !important',
                "margin": '4px !important'
            },
            ".ag-theme-streamlit [col-id='Elimina'] .ag-cell:hover": {
                "background": 'linear-gradient(135deg, rgba(220, 20, 60, 0.25) 0%, rgba(139, 0, 0, 0.3) 100%) !important',
                "border-color": 'rgba(220, 20, 60, 0.7) !important',
                "transform": 'scale(1.05) !important'
            },
            # Styling checkbox nella colonna Elimina
            ".ag-theme-streamlit [col-id='Elimina'] input[type='checkbox']": {
                "width": '20px !important',
                "height": '20px !important',
                "accent-color": '#ff4757 !important',
                "cursor": 'pointer !important'
            },
            # CSS specifico per colonne di testo con text wrapping ottimizzato
            ".ag-theme-streamlit [col-id='Descrizione'] .ag-cell": {
                "text-align": 'left !important',
                "justify-content": 'flex-start !important',
                "align-items": 'flex-start !important',
                "padding": '8px !important',
                "white-space": 'normal !important',
                "word-wrap": 'break-word !important',
                "word-break": 'break-word !important',
                "overflow-wrap": 'break-word !important',
                "line-height": '1.4 !important',
                "display": 'block !important',
                "height": 'auto !important',
                "min-height": '60px !important',
                "max-width": '350px !important',
                "hyphens": 'auto !important'
            },
            ".ag-theme-streamlit [col-id='Contromisura'] .ag-cell": {
                "text-align": 'left !important',
                "justify-content": 'flex-start !important',
                "align-items": 'flex-start !important',
                "padding": '8px !important',
                "white-space": 'normal !important',
                "word-wrap": 'break-word !important',
                "word-break": 'break-word !important',
                "overflow-wrap": 'break-word !important',
                "line-height": '1.4 !important',
                "display": 'block !important',
                "height": 'auto !important',
                "min-height": '60px !important',
                "max-width": '350px !important',
                "hyphens": 'auto !important'
            },
            # CSS per controllo wrapping colonne numeriche e di controllo
            ".ag-theme-streamlit [col-id='ID'] .ag-cell": {
                "max-width": '80px !important',
                "word-wrap": 'break-word !important'
            },
            ".ag-theme-streamlit [col-id='Probabilità'] .ag-cell": {
                "max-width": '120px !important',
                "word-wrap": 'break-word !important'
            },
            ".ag-theme-streamlit [col-id='Impatto'] .ag-cell": {
                "max-width": '120px !important',
                "word-wrap": 'break-word !important'
            },
            ".ag-theme-streamlit [col-id='Priorità'] .ag-cell": {
                "max-width": '150px !important',
                "word-wrap": 'break-word !important'
            },
            ".ag-theme-streamlit [col-id='Stato'] .ag-cell": {
                "max-width": '100px !important',
                "word-wrap": 'break-word !important'
            },
            ".ag-theme-streamlit [col-id='Data scadenza'] .ag-cell": {
                "max-width": '130px !important',
                "word-wrap": 'break-word !important'
            },
            ".ag-theme-streamlit [col-id='Elimina'] .ag-cell": {
                "max-width": '100px !important'
            }
        }
    )

    # ===========================
    # GESTIONE ELIMINAZIONE RISCHI
    # ===========================

    # Elaborazione response griglia per gestire eliminazioni
    if 'data' in grid_response and grid_response['data'] is not None:
        updated_df = pd.DataFrame(grid_response['data'])
        
        # Ricerca righe marcate per eliminazione tramite checkbox
        if 'Elimina' in updated_df.columns:
            rows_to_delete = updated_df[updated_df['Elimina'] == True]
            
            if not rows_to_delete.empty:
                # Estrazione ID rischi da eliminare
                ids_to_delete = rows_to_delete['ID'].tolist()
                
                # Rimozione rischi dal dataset principale
                for risk_id in ids_to_delete:
                    st.session_state.df = st.session_state.df[st.session_state.df['ID'] != risk_id]
                
                # Persistenza immediata e refresh interfaccia
                if save_data(st.session_state.df, DATA_FILE):
                    st.rerun()
                else:
                    st.error(f"Errore nel salvataggio dopo eliminazione")
else:
    # Messaggio informativo quando non ci sono rischi
    st.info("Nessun rischio presente.")

# ===========================
# HEAT MAP INTERATTIVA
# ===========================

# Visualizzazione heat map solo se ci sono dati
if not st.session_state.df.empty:
    st.header("Heat Map dei Rischi")

    # Definizione etichette assi per griglia 5x5
    impact_labels = ['1', '2', '3', '4', '5']
    likelihood_labels = ['1', '2', '3', '4', '5']
    
    def get_risk_color_and_priority(prob, imp):
        """
        Determina colore e priorità basato sul valore del rischio.
        
        Args:
            prob (float): Valore probabilità (1-5)
            imp (float): Valore impatto (1-5)
            
        Returns:
            tuple: (colore_hex, etichetta_priorità)
        """
        risk_value = prob * imp
        if risk_value >= 16:
            return '#ef4444', 'ESTREMA'  # Rosso per rischio critico
        elif risk_value >= 11:
            return '#f97316', 'ALTA'     # Arancione per rischio alto
        elif risk_value >= 6:
            return '#eab308', 'MEDIA'    # Giallo per rischio medio
        elif risk_value >= 1:
            return '#22c55e', 'BASSA'    # Verde per rischio basso
        else:
            return '#374151', 'NESSUNO'  # Grigio per nessun rischio
    
    # ===========================
    # GENERAZIONE HTML HEAT MAP
    # ===========================
    
    # Container principale con styling premium
    heatmap_complete = '<div style="display: flex; justify-content: center; margin: 20px 0; padding: 15px;">'
    heatmap_complete += '<div style="background: linear-gradient(135deg, #1e293b 0%, #334155 100%); padding: 40px; border-radius: 16px; box-shadow: 0 15px 40px rgba(0,0,0,0.4); border: 1px solid #475569; max-width: 800px; width: 100%; position: relative;">'
    
    # Etichetta IMPATTO verticale posizionata precisamente
    heatmap_complete += '<div style="position: absolute; left: 75px; top: 290px; transform: translateY(-50%) rotate(-90deg); font-weight: 700; color: #e2e8f0; font-size: 16px; text-transform: uppercase; letter-spacing: 0.8px; text-shadow: 0 1px 2px rgba(0,0,0,0.5);">IMPATTO</div>'
    
    # Container griglia con dimensioni ottimizzate
    heatmap_complete += '<div style="position: relative; width: 520px; height: 520px; margin: 0 auto;">'
    
    # Parametri griglia 5x5 con dimensioni precise
    grid_width = 450  
    grid_height = 450  
    cell_width = grid_width / 5
    cell_height = grid_height / 5
    
    # Area contenitore griglia di sfondo
    heatmap_complete += f'<div style="position: absolute; top: 29px; left: 80px; width: {grid_width}px; height: {grid_height}px;">'
    
    # Generazione celle di sfondo con colori basati su priorità
    for i in range(5):  # Righe: Impatto da alto (top) a basso (bottom)
        for j in range(5):  # Colonne: Probabilità da bassa (left) a alta (right)
            # Calcolo valori per cella della griglia
            prob = j + 1  # Probabilità: 1-5 da sinistra a destra
            imp = 5 - i   # Impatto: 5-1 da alto a basso (inversione per visualizzazione)
            color, priority = get_risk_color_and_priority(prob, imp)
            
            # Posizionamento cella
            x = j * cell_width
            y = i * cell_height
            
            # Creazione cella con tooltip informativo
            heatmap_complete += f'<div style="position: absolute; left: {x}px; top: {y}px; width: {cell_width}px; height: {cell_height}px; background-color: {color}; border: 2px solid #334155; border-radius: 8px; box-shadow: 0 3px 8px rgba(0,0,0,0.2);" title="Prob: {prob}, Imp: {imp}, Priorità: {priority}"></div>'
    
    # ===========================
    # POSIZIONAMENTO RISCHI SULLA GRIGLIA
    # ===========================
    
    # Raggruppamento rischi per posizione per gestire sovrapposizioni
    risk_positions = {}
    for _, row in st.session_state.df.iterrows():
        prob = float(row['Probabilità'])
        imp = float(row['Impatto'])
        risk_id = int(row['ID'])
        position_key = (prob, imp)
        
        if position_key not in risk_positions:
            risk_positions[position_key] = []
        risk_positions[position_key].append(risk_id)
    
    # Posizionamento marcatori rischi con coordinate precise
    for (prob, imp), risk_ids in risk_positions.items():
        # Validazione range valori per sicurezza
        if prob < 1 or prob > 5 or imp < 1 or imp > 5:
            continue  
            
        risk_value = prob * imp
        
        # Calcolo coordinate pixel precise basate su valori decimali
        pixel_x = ((prob - 0.5) * cell_width)
        pixel_y = ((5.5 - imp) * cell_height)
        
        # Dimensionamento dinamico basato su numero di rischi sovrapposti
        num_risks = len(risk_ids)
        size = 28 + (num_risks - 1) * 6
        font_size = 11 if num_risks == 1 else 10
        
        # Creazione stringa ID per etichetta
        id_string = ", ".join(map(str, sorted(risk_ids)))
        
        # Posizionamento finale centrato
        final_x = round(pixel_x - size/2)
        final_y = round(pixel_y - size/2)
        
        # Creazione marcatore visuale con tooltip dettagliato
        heatmap_complete += f'<div style="position: absolute; left: {final_x}px; top: {final_y}px; width: {size}px; height: {size}px; background: #000000; color: #ffffff; border-radius: 50%; display: flex; align-items: center; justify-content: center; font-weight: 700; box-shadow: 0 3px 8px rgba(0,0,0,0.6); border: 2px solid #ffffff; font-size: {font_size}px; z-index: 10;" title="Rischi ID: {id_string} - Prob: {prob}, Imp: {imp}, Valore: {risk_value:.1f}">{id_string}</div>'
    
    heatmap_complete += '</div>'  # Chiusura area griglia
    
    # ===========================
    # ETICHETTATURA ASSI
    # ===========================
    
    # Etichette asse Y (Impatto) - posizionate verticalmente a sinistra
    for i, label in enumerate(reversed(impact_labels)):
        y_pos = 25 + i * cell_height + cell_height/2 - 10
        heatmap_complete += f'<div style="position: absolute; left: 35px; top: {y_pos}px; font-size: 12px; font-weight: 600; color: #cbd5e1; text-transform: uppercase; letter-spacing: 0.3px; width: 40px; text-align: right;">{label}</div>'
    
    # Etichette asse X (Probabilità) - posizionate orizzontalmente in basso
    for i, label in enumerate(likelihood_labels):
        x_pos = 80 + i * cell_width + cell_width/2 - 10
        heatmap_complete += f'<div style="position: absolute; left: {x_pos}px; top: {25 + grid_height + 15}px; font-size: 12px; font-weight: 600; color: #cbd5e1; text-transform: uppercase; letter-spacing: 0.3px; width: 20px; text-align: center;">{label}</div>'

    # Etichetta asse X principale centrata
    heatmap_complete += '<div style="position: absolute; left: 305px; top: 520px; transform: translateX(-50%); font-weight: 700; color: #e2e8f0; font-size: 16px; text-transform: uppercase; letter-spacing: 0.8px; text-shadow: 0 1px 2px rgba(0,0,0,0.5);">PROBABILITÀ</div>'
    
    # Chiusura container principali
    heatmap_complete += '</div>'  # Chiusura contenitore griglia
    heatmap_complete += '</div>'  # Chiusura wrapper heat map
    heatmap_complete += '</div>'  # Chiusura container principale
    
    # Rendering HTML heat map
    st.markdown(heatmap_complete, unsafe_allow_html=True)

# ===========================
# SEZIONE ESPORTAZIONE DATI
# ===========================

# Interfaccia esportazione disponibile solo con dati presenti
if not st.session_state.df.empty:
    st.header("Esportazione Dati")
    
    # Spaziatura visuale
    st.markdown("<div style='margin-top: 20px;'></div>", unsafe_allow_html=True)
    
    # Layout a due colonne per i pulsanti di esportazione
    col1, col2 = st.columns(2)
    
    # ===========================
    # ESPORTAZIONE PDF
    # ===========================
    
    with col1:
        if st.button("📄 Esporta in PDF", use_container_width=True):
            with st.spinner("Generazione PDF in corso..."):
                pdf_buffer = create_pdf_report(st.session_state.df)
                
                if pdf_buffer is not None:
                    # Download button con filename dinamico
                    st.download_button(
                        label="⬇️ Scarica Report PDF",
                        data=pdf_buffer,
                        file_name=f"risk_assessment_report_{date.today()}.pdf",
                        mime="application/pdf",
                        use_container_width=True
                    )
                    st.success("PDF generato con successo!")
                else:
                    st.error("Impossibile generare il PDF. Verifica che la libreria ReportLab sia installata.")
    
    # ===========================
    # ESPORTAZIONE EXCEL
    # ===========================
    
    with col2:
        if st.button("📊 Esporta in Excel", use_container_width=True):
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.utils.dataframe import dataframe_to_rows
                
                # Creazione workbook con foglio principale
                wb = Workbook()
                ws = wb.active
                ws.title = "Risk Assessment"
                
                # ===========================
                # DEFINIZIONE STILI EXCEL
                # ===========================
                
                # Stili per header tabella
                header_font = Font(bold=True, color="FFFFFF", size=12)
                header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                # Bordi per celle
                border = Border(
                    left=Side(style='thin', color='334155'),
                    right=Side(style='thin', color='334155'),
                    top=Side(style='thin', color='334155'),
                    bottom=Side(style='thin', color='334155')
                )
                
                # ===========================
                # INTESTAZIONE DOCUMENTO
                # ===========================
                
                # Titolo principale con merge celle
                ws.merge_cells('A1:H1')
                ws['A1'] = '🛡️ RISK ASSESSMENT REPORT'
                ws['A1'].font = Font(bold=True, size=20, color="DC143C")
                ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
                ws.row_dimensions[1].height = 40
                
                # Data generazione report
                ws.merge_cells('A2:H2')
                ws['A2'] = f'Data Report: {date.today().strftime("%d/%m/%Y")}'
                ws['A2'].font = Font(size=12, italic=True)
                ws['A2'].alignment = Alignment(horizontal="center", vertical="center")
                ws.row_dimensions[2].height = 25
                
                # Riga vuota per spaziatura
                ws.row_dimensions[3].height = 10
                
                # ===========================
                # TABELLA DATI PRINCIPALE
                # ===========================
                
                # Header tabella con ordine colonne identico a AgGrid (escluso Valore Rischio aggiuntivo)
                headers = ['ID', 'Descrizione', 'Probabilità', 'Impatto', 'Valore Rischio', 'Priorità', 'Contromisura', 'Stato', 'Data Scadenza']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=4, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border
                
                # Popolamento dati con ordine colonne coerente con AgGrid e text wrapping abilitato
                for row_idx, (_, row) in enumerate(st.session_state.df.iterrows(), start=5):
                    # Colonna 1: ID
                    ws.cell(row=row_idx, column=1, value=row['ID']).border = border
                    
                    # Colonna 2: Descrizione con text wrapping
                    desc_cell = ws.cell(row=row_idx, column=2, value=row['Descrizione'])
                    desc_cell.border = border
                    desc_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                    
                    # Colonne 3-5: Valori numerici
                    ws.cell(row=row_idx, column=3, value=f"{row['Probabilità']:.1f}").border = border
                    ws.cell(row=row_idx, column=4, value=f"{row['Impatto']:.1f}").border = border
                    ws.cell(row=row_idx, column=5, value=f"{row['Valore_Rischio']:.1f}").border = border
                    
                    # Colonna 6: Priorità con colore condizionale
                    priority_cell = ws.cell(row=row_idx, column=6, value=row['Priorità'])
                    priority_cell.border = border
                    priority_cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Applicazione colori per priorità
                    if row['Priorità'] == 'Estrema':
                        priority_cell.fill = PatternFill(start_color="FFE4E1", end_color="FFE4E1", fill_type="solid")
                        priority_cell.font = Font(bold=True, color="8B0000")
                    elif row['Priorità'] == 'Alta':
                        priority_cell.fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")
                        priority_cell.font = Font(bold=True, color="FF8C00")
                    elif row['Priorità'] == 'Media':
                        priority_cell.fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
                        priority_cell.font = Font(bold=True, color="FFD700")
                    else:
                        priority_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                        priority_cell.font = Font(bold=True, color="006400")
                    
                    # Colonna 7: Contromisura con text wrapping
                    cont_cell = ws.cell(row=row_idx, column=7, value=row['Contromisura'])
                    cont_cell.border = border
                    cont_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                    
                    # Colonna 8: Stato
                    ws.cell(row=row_idx, column=8, value=row['Stato']).border = border
                    
                    # Colonna 9: Data Scadenza
                    ws.cell(row=row_idx, column=9, value=row['Data scadenza']).border = border
                    
                    # Allineamento centrale per colonne numeriche e di controllo
                    for col in [1, 3, 4, 5, 6, 8, 9]:
                        ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Impostazione altezza riga per accommodare text wrapping su più righe
                    ws.row_dimensions[row_idx].height = 60
                
                # ===========================
                # AGGIUNTA HEAT MAP AL FOGLIO PRINCIPALE (gestione migliorata)
                # ===========================
                
                # Generazione heat map come immagine per Excel
                heatmap_img_buffer = create_heatmap_image(st.session_state.df)
                if heatmap_img_buffer is not None:
                    try:
                        # Importazione PIL per gestione immagini in Excel
                        from PIL import Image as PILImage
                        from openpyxl.drawing.image import Image as ExcelImage
                        import io
                        
                        # Aggiunta spazio e titolo per heat map
                        current_row = len(st.session_state.df) + 7  # Dopo dati + spazio
                        
                        # Titolo heat map
                        ws.merge_cells(f'A{current_row}:I{current_row}')
                        title_cell = ws[f'A{current_row}']
                        title_cell.value = 'HEAT MAP DEI RISCHI'
                        title_cell.font = Font(bold=True, size=16, color="DC143C")
                        title_cell.alignment = Alignment(horizontal="center", vertical="center")
                        ws.row_dimensions[current_row].height = 30
                        
                        # Creazione immagine direttamente da buffer (senza file temporaneo)
                        img = ExcelImage(heatmap_img_buffer)
                        img.width = 600   # Larghezza ottimizzata per Excel
                        img.height = 400  # Altezza proporzionale
                        
                        # Posizionamento immagine
                        ws.add_image(img, f'B{current_row + 2}')
                        
                        # Note esplicative sotto la heat map
                        note_row = current_row + 22  # Spazio per immagine + margine
                        ws.merge_cells(f'A{note_row}:I{note_row}')
                        note_cell = ws[f'A{note_row}']
                        note_cell.value = 'Note: I numeri neri indicano gli ID dei rischi posizionati secondo probabilità e impatto (valori da 1 a 5 con incrementi di 0.5). I colori rappresentano le priorità: Verde (Bassa), Giallo (Media), Arancione (Alta), Rosso (Estrema).'
                        note_cell.font = Font(size=10, italic=True)
                        note_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                        ws.row_dimensions[note_row].height = 40
                        
                    except ImportError:
                        # Aggiunta nota se PIL non è disponibile
                        note_row = len(st.session_state.df) + 8
                        ws.merge_cells(f'A{note_row}:I{note_row}')
                        note_cell = ws[f'A{note_row}']
                        note_cell.value = 'Heat Map non disponibile: installare Pillow con "pip install Pillow" per includere visualizzazioni grafiche.'
                        note_cell.font = Font(size=12, italic=True, color="FF0000")
                        note_cell.alignment = Alignment(horizontal="center", vertical="center")
                    except Exception as e:
                        # Gestione errori generici con messaggio migliorato
                        note_row = len(st.session_state.df) + 8
                        ws.merge_cells(f'A{note_row}:I{note_row}')
                        note_cell = ws[f'A{note_row}']
                        note_cell.value = 'Heat Map non disponibile: errore nella generazione della visualizzazione.'
                        note_cell.font = Font(size=12, italic=True, color="FF0000")
                        note_cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # ===========================
                # LARGHEZZE COLONNE OTTIMIZZATE (ridotte per eliminare spazio vuoto)
                # ===========================
                
                ws.column_dimensions['A'].width = 8    # ID - compatta (coerente con 60px AgGrid)
                ws.column_dimensions['B'].width = 45   # Descrizione (coerente con 300px AgGrid)
                ws.column_dimensions['C'].width = 12   # Probabilità (coerente con 100px AgGrid)
                ws.column_dimensions['D'].width = 10   # Impatto (coerente con 90px AgGrid)
                ws.column_dimensions['E'].width = 15   # Valore Rischio - calcolato
                ws.column_dimensions['F'].width = 12   # Priorità (coerente con 100px AgGrid)
                ws.column_dimensions['G'].width = 40   # Contromisura (coerente con 280px AgGrid)
                ws.column_dimensions['H'].width = 12   # Stato (coerente con 80px AgGrid)
                ws.column_dimensions['I'].width = 15   # Data Scadenza (coerente con 110px AgGrid)
                
                # ===========================
                # FOGLIO RIEPILOGO STATISTICHE
                # ===========================
                
                # Creazione secondo foglio per statistiche
                ws2 = wb.create_sheet("Riepilogo")
                
                # Titolo riepilogo
                ws2.merge_cells('A1:B1')
                ws2['A1'] = 'RIEPILOGO RISCHI PER PRIORITÀ'
                ws2['A1'].font = Font(bold=True, size=16, color="DC143C")
                ws2['A1'].alignment = Alignment(horizontal="center", vertical="center")
                
                # Calcolo statistiche per priorità
                priority_counts = st.session_state.df['Priorità'].value_counts()
                
                # Header tabella riepilogo
                ws2['A3'] = 'Priorità'
                ws2['B3'] = 'Numero Rischi'
                ws2['A3'].font = header_font
                ws2['A3'].fill = header_fill
                ws2['B3'].font = header_font
                ws2['B3'].fill = header_fill
                
                # Popolamento dati riepilogo in ordine di priorità
                row = 4
                for priority in ['Estrema', 'Alta', 'Media', 'Bassa']:
                    ws2[f'A{row}'] = priority
                    ws2[f'B{row}'] = priority_counts.get(priority, 0)
                    row += 1
                
                # Riga totale
                ws2[f'A{row}'] = 'TOTALE'
                ws2[f'B{row}'] = len(st.session_state.df)
                ws2[f'A{row}'].font = Font(bold=True)
                ws2[f'B{row}'].font = Font(bold=True)
                
                # Ottimizzazione larghezze foglio riepilogo
                ws2.column_dimensions['A'].width = 15
                ws2.column_dimensions['B'].width = 15
                
                # ===========================
                # GENERAZIONE FILE FINALE
                # ===========================
                
                # Salvataggio in buffer per download
                excel_buffer = BytesIO()
                wb.save(excel_buffer)
                excel_buffer.seek(0)
                
                # Download button con filename dinamico
                st.download_button(
                    label="⬇️ Scarica Report Excel",
                    data=excel_buffer,
                    file_name=f"risk_assessment_report_{date.today()}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                st.success("Excel generato con successo!")
                
            except ImportError:
                st.error("Libreria openpyxl non disponibile. Installa con: pip install openpyxl")
            except Exception as e:
                st.error(f"Errore nella generazione Excel: {str(e)}")

# ===========================
# FOOTER E INFORMAZIONI
# ===========================

# Footer informativo (opzionale)
st.markdown("---")
st.markdown(
    "<div style='text-align: center; color: #64748b; font-size: 0.9em;'>"
    "Dashboard Risk Assessment - Gestione professionale dei rischi di progetto"
    "</div>", 
    unsafe_allow_html=True
)